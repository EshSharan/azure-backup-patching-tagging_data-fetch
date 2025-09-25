import json
import logging
from azure.identity import ClientSecretCredential
from azure.mgmt.subscription import SubscriptionClient
from azure.mgmt.recoveryservicesbackup import RecoveryServicesBackupClient
from azure.mgmt.recoveryservices import RecoveryServicesClient
from azure.mgmt.compute import ComputeManagementClient
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
logging.basicConfig(
   level=logging.INFO,
   format='%(asctime)s - %(levelname)s - %(message)s',
   handlers=[logging.FileHandler("backup_report_log.txt", encoding='utf-8')]
)
def normalize_id(rid: str) -> str:
   return rid.lower().replace(" ", "") if rid else ""
def get_vm_info(compute_client, resource_group, resource_name):
   os_type, os_version, status = "-", "-", "-"
   try:
       vm = compute_client.virtual_machines.get(resource_group, resource_name, expand='instanceView')
       if vm.storage_profile and vm.storage_profile.os_disk:
           os_type = vm.storage_profile.os_disk.os_type
       if vm.instance_view:
           os_name = getattr(vm.instance_view, 'os_name', None)
           os_ver = getattr(vm.instance_view, 'os_version', None)
           if os_name and os_ver:
               os_version = f"{os_name} {os_ver}"
           statuses = getattr(vm.instance_view, 'statuses', [])
           for s in statuses:
               code = getattr(s, 'code', '')
               if code.startswith('PowerState/'):
                   status = code.split('/')[-1]
   except Exception as e:
       logging.warning(f"Could not get VM info for '{resource_name}': {e}")
   return os_type, os_version, status
def run_backup():
   config_file = "conf.json"
   try:
       with open(config_file, "r") as f:
           config = json.load(f)
           tenant_id = config.get("tenantId")
           client_id = config.get("clientId")
           client_secret = config.get("clientSecret")
   except Exception as e:
       logging.critical(f"Failed to read '{config_file}': {e}")
       return
   credential = ClientSecretCredential(
       tenant_id=tenant_id,
       client_id=client_id,
       client_secret=client_secret
   )
   subscription_client = SubscriptionClient(credential)
   wb = Workbook()
   backup_sheet = wb.active
   backup_sheet.title = "Backup Sheet"
   BACKUP_HEADERS = [
       "Name of the Resource", "Subscription", "Resource Group", "VM Status", "Operating System Type", "Operating System Version",
       "RSV", "Policy Name", "Policy Tier", "Schedule", "Instant Snapshot", "Daily Retention",
       "Weekly Retention", "Monthly Retention", "Yearly Retention", "Tiering"
   ]
   backup_sheet.append(BACKUP_HEADERS)
   logging.info("Starting backup collection... 💾")
   for subscription in subscription_client.subscriptions.list():
       subscription_name = subscription.display_name
       print(f"Processing subscription: {subscription_name}")
       recovery_services_client = RecoveryServicesClient(credential, subscription.subscription_id)
       backup_client = RecoveryServicesBackupClient(credential, subscription.subscription_id)
       compute_client = ComputeManagementClient(credential, subscription.subscription_id)
       backup_info_cache = {}
       try:
           vaults = recovery_services_client.vaults.list_by_subscription_id()
           for vault in vaults:
               try:
                   resource_group = vault.id.split('/')[4]
                   protected_items = backup_client.backup_protected_items.list(
                       resource_group_name=resource_group,
                       vault_name=vault.name,
                       filter="backupManagementType eq 'AzureIaasVM'"
                   )
                   for protected_item in protected_items:
                       try:
                           source_resource_id = getattr(protected_item.properties, 'source_resource_id', None)
                           if not source_resource_id:
                               continue
                           policy_id = getattr(protected_item.properties, 'policy_id', None)
                           policy_name = policy_id.split('/')[-1] if policy_id else None
                           if not policy_name:
                               continue
                           policy = backup_client.protection_policies.get(
                               vault_name=vault.name,
                               resource_group_name=resource_group,
                               policy_name=policy_name
                           )
                           policy_tier = "Standard"
                           schedule = "-"
                           instant_snapshot = "-"
                           daily_retention = "-"
                           weekly_retention = "-"
                           monthly_retention = "-"
                           yearly_retention = "-"
                           tiering = "No"
                           policy_props = policy.properties
                           if getattr(policy_props, 'instant_recovery_policy', None):
                               policy_tier = "Enhanced"
                           elif policy_name and "Enhanced" in policy_name:
                               policy_tier = "Enhanced"
                           schedule_policy = getattr(policy_props, 'schedule_policy', None)
                           if schedule_policy:
                               freq = getattr(schedule_policy, 'schedule_run_frequency', None)
                               if freq == "Daily" and hasattr(schedule_policy, 'schedule_run_times') and schedule_policy.schedule_run_times:
                                   time_val = schedule_policy.schedule_run_times[0].strftime("%H:%M UTC")
                                   schedule = f"Daily at {time_val}"
                               elif freq == "Weekly":
                                   days = getattr(schedule_policy, 'schedule_days_of_week', [])
                                   times = getattr(schedule_policy, 'schedule_run_times', [])
                                   if days and times:
                                       time_val = times[0].strftime("%H:%M UTC")
                                       days_str = ", ".join(days)
                                       schedule = f"Weekly on {days_str} at {time_val}"
                                   elif days:
                                       days_str = ", ".join(days)
                                       schedule = f"Weekly on {days_str}"
                                   elif times:
                                       time_val = times[0].strftime("%H:%M UTC")
                                       schedule = f"Weekly at {time_val}"
                                   else:
                                       schedule = "Weekly"
                               else:
                                   schedule = freq or "-"
                           instant_snapshot_days = getattr(policy_props, 'instant_rp_retention_range_in_days', None)
                           if instant_snapshot_days is not None:
                               instant_snapshot = f"{instant_snapshot_days} Days"
                           retention_policy = getattr(policy_props, 'retention_policy', None)
                           if retention_policy:
                               if hasattr(retention_policy, 'daily_schedule') and retention_policy.daily_schedule and retention_policy.daily_schedule.retention_duration:
                                   daily_retention = f"{retention_policy.daily_schedule.retention_duration.count} Days"
                               if hasattr(retention_policy, 'weekly_schedule') and retention_policy.weekly_schedule and retention_policy.weekly_schedule.retention_duration:
                                   weekly_retention = f"{retention_policy.weekly_schedule.retention_duration.count} Weeks"
                               if hasattr(retention_policy, 'monthly_schedule') and retention_policy.monthly_schedule and retention_policy.monthly_schedule.retention_duration:
                                   monthly_retention = f"{retention_policy.monthly_schedule.retention_duration.count} Months"
                               if hasattr(retention_policy, 'yearly_schedule') and retention_policy.yearly_schedule and retention_policy.yearly_schedule.retention_duration:
                                   yearly_retention = f"{retention_policy.yearly_schedule.retention_duration.count} Years"
                           if getattr(policy_props, 'tiering_policy', None):
                               tiering = "Yes"
                           backup_info_cache[normalize_id(source_resource_id)] = [
                               vault.name, policy_name, policy_tier, schedule, instant_snapshot,
                               daily_retention, weekly_retention, monthly_retention, yearly_retention, tiering
                           ]
                       except Exception as inner_e:
                           logging.error(f"Error processing protected item '{protected_item.name}': {inner_e}")
                           continue
               except Exception as vault_e:
                   logging.error(f"Error processing vault {vault.name}: {vault_e}")
                   continue
       except Exception as e:
           logging.error(f"Error listing vaults for subscription {subscription_name}: {e}")
       for res_id, info in backup_info_cache.items():
           try:
               parts = res_id.split('/')
               resource_group = parts[4] if len(parts) > 4 else "-"
               resource_name = parts[-1] if len(parts) > 0 else "-"
           except Exception:
               resource_group = "-"
               resource_name = "-"
           os_type, os_version, status = get_vm_info(compute_client, resource_group, resource_name)
           backup_sheet.append([resource_name, subscription_name, resource_group, status, os_type, os_version, *info])
   output_file = "backup_report.xlsx"
   wb.save(output_file)
   # --- Styling with openpyxl ---
   wb = load_workbook(output_file)
   ws = wb.active
   thin_border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
   header_fill = PatternFill(start_color="FF00FF", end_color="FF00FF", fill_type="solid")
   header_font = Font(color="FFFFFF", bold=True, name="Calibri", size=12)
   for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=ws.max_column):
       for cell in col:
           cell.fill = header_fill
           cell.font = header_font
           cell.border = thin_border
           cell.alignment = Alignment(horizontal='center', vertical='center')
   data_font = Font(color="000000", name="Calibri", size=12)
   data_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
   for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
       for cell in row:
           cell.font = data_font
           cell.fill = data_fill
           cell.border = thin_border
   wb.save(output_file)
   print(f"\n✅ Backup report saved and formatted successfully as '{output_file}' 🎉")
if __name__ == "__main__":
   run_backup()