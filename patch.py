import pandas as pd
import json
import time
import sys
from datetime import datetime, timedelta
from azure.identity import ClientSecretCredential
from azure.mgmt.subscription import SubscriptionClient
from azure.mgmt.compute import ComputeManagementClient
from azure.mgmt.maintenance import MaintenanceManagementClient
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
import os
def get_week_order(patch_day_str):
   week_map = {"First": 1, "Second": 2, "Third": 3, "Fourth": 4, "Last": 5}
   for k, v in week_map.items():
       if patch_day_str.startswith(k):
           return v
   return 99
def get_azure_update_manager_data(config_file: str):
   try:
       with open(config_file, 'r') as f:
           config = json.load(f)
       credential = ClientSecretCredential(
           tenant_id=config['tenantId'],
           client_id=config['clientId'],
           client_secret=config['clientSecret']
       )
   except Exception as e:
       print(f"Authentication failed: {e}")
       sys.exit(1)
   subscription_client = SubscriptionClient(credential=credential)
   try:
       subscriptions = list(subscription_client.subscriptions.list())
   except Exception as e:
       print(f"Failed to list subscriptions: {e}")
       sys.exit(1)
   all_vm_data = []
   os.makedirs("debug_raw_configs", exist_ok=True)
   for sub in subscriptions:
       sub_name = sub.display_name
       sub_id = sub.subscription_id
       compute_client = ComputeManagementClient(credential, sub_id)
       maintenance_client = MaintenanceManagementClient(credential, sub_id)
       try:
           vms = list(compute_client.virtual_machines.list_all())
           if not vms:
               continue
       except Exception:
           continue
       for vm in vms:
           vm_name = vm.name
           resource_group_name = vm.id.split('/')[4]
           # OS info
           os_type = getattr(vm.storage_profile.os_disk, "os_type", "-")
           os_version = getattr(vm.storage_profile.image_reference, "offer", "-") if getattr(vm.storage_profile, "image_reference", None) else "-"
           vm_data = {
               "Subscription Name": sub_name,
               "Resource Group": resource_group_name,
               "VM Location": vm.location,
               "VM Name": vm_name,
               "Operating System": os_type,
               "OS Version": os_version,
               "Update Manager Status": "Not configured",
               "Maintenance Schedule Name": "-",
               "Patching Day": "-",
               "Maintenance Window Duration": "-",
               "Patching Downtime": "-",
               "Time Zone": "-",
               "Validation Time": "-",
               "Batch": "-",
               "Reboot Setting": "-",
               "_start_datetime": None
           }
           try:
               assignments = list(maintenance_client.configuration_assignments.list_parent(
                   resource_group_name=resource_group_name,
                   provider_name="Microsoft.Compute",
                   resource_parent_type="",
                   resource_parent_name="",
                   resource_type="virtualMachines",
                   resource_name=vm_name
               ))
               if assignments:
                   vm_data["Update Manager Status"] = "Configured with schedule"
                   assignment = assignments[0]
                   mc_parts = assignment.maintenance_configuration_id.split('/')
                   mc_resource_group = mc_parts[4]
                   mc_name = mc_parts[-1]
                   mc_details = maintenance_client.maintenance_configurations.get(mc_resource_group, mc_name)
                   vm_data["Maintenance Schedule Name"] = mc_name
                   mc_dict = mc_details.as_dict()
                   debug_file = os.path.join("debug_raw_configs", f"{vm_name}_config.json")
                   with open(debug_file, "w") as df:
                       json.dump(mc_dict, df, indent=2)
                   start_dt = mc_dict.get("start_date_time")
                   duration = mc_dict.get("duration")
                   recur_every = mc_dict.get("recur_every")
                   time_zone = mc_dict.get("time_zone")
                   install_patches = mc_dict.get("install_patches", {})
                   reboot_setting = install_patches.get("reboot_setting", "-")
                   start_time_str = None
                   dt_obj = None
                   if start_dt:
                       try:
                           dt_obj = datetime.strptime(start_dt, "%Y-%m-%d %H:%M")
                           start_time_str = dt_obj.strftime("%I:%M %p")
                           vm_data["_start_datetime"] = dt_obj
                       except Exception:
                           pass
                   if recur_every:
                       parts = recur_every.split()
                       if len(parts) >= 3:
                           vm_data["Patching Day"] = f"{parts[1]} {parts[2]} of the month"
                       else:
                           vm_data["Patching Day"] = recur_every
                   hours_str = "-"
                   duration_hours = None
                   if duration:
                       try:
                           hrs, mins = duration.split(":")
                           hrs, mins = int(hrs), int(mins)
                           duration_hours = hrs + mins / 60
                           hours_str = (f"{hrs} hours " if hrs else "") + (f"{mins} mins" if mins else "")
                       except Exception:
                           pass
                   vm_data["Maintenance Window Duration"] = hours_str or "-"
                   vm_data["Time Zone"] = time_zone or "-"
                   if start_time_str and duration_hours is not None and dt_obj:
                       try:
                           end_time = dt_obj + timedelta(hours=duration_hours)
                           end_time_str = end_time.strftime("%I:%M %p")
                           vm_data["Patching Downtime"] = f"{start_time_str} - {end_time_str} {time_zone}"
                       except Exception:
                           pass
                   vm_data["Reboot Setting"] = reboot_setting
           except Exception:
               pass
           all_vm_data.append(vm_data)
           time.sleep(0.2)
   if all_vm_data:
       df = pd.DataFrame(all_vm_data)
       # Assign batches
       schedule_start_map = {}
       for _, row in df.iterrows():
           sched = row["Maintenance Schedule Name"]
           start_dt = row["_start_datetime"]
           if sched != "-" and sched not in schedule_start_map:
               schedule_start_map[sched] = start_dt
       schedule_order_list = sorted(
           schedule_start_map.keys(),
           key=lambda x: (get_week_order(df[df["Maintenance Schedule Name"] == x]["Patching Day"].iloc[0]),
                          schedule_start_map[x])
       )
       batch_map = {name: f"Batch {i+1}" for i, name in enumerate(schedule_order_list)}
       df["Batch"] = df["Maintenance Schedule Name"].map(batch_map).fillna("-")
       df.drop(columns=["_start_datetime"], inplace=True)
       # Reorder columns: Subscription Name, Resource Group, VM Location, VM Name, OS, OS Version, ...
       cols = [
           "Subscription Name", "Resource Group", "VM Location",
           "VM Name", "Operating System", "OS Version",
           "Update Manager Status", "Maintenance Schedule Name", "Patching Day",
           "Maintenance Window Duration", "Patching Downtime", "Time Zone",
           "Validation Time", "Batch", "Reboot Setting"
       ]
       df = df[cols]
       file_name = "Patching-Sheet.xlsx"
       df.to_excel(file_name, index=False)
       # --- Styling with openpyxl ---
       wb = load_workbook(file_name)
       ws = wb.active
       thin_border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))
       header_fill = PatternFill(start_color="FF00FF", end_color="FF00FF", fill_type="solid")
       header_font = Font(color="FFFFFF", bold=True, name="Calibri", size=12)
       for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=ws.max_column):
           for cell in col:
               cell.fill = header_fill
               cell.font = header_font
               cell.border = thin_border
               cell.alignment = Alignment(horizontal='center', vertical='center')
       data_font = Font(color="000000", name="Calibri", size=12)
       data_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
       for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
           for cell in row:
               cell.font = data_font
               cell.fill = data_fill
               cell.border = thin_border
       wb.save(file_name)
       print(f"\n✅ Data exported and formatted successfully to {file_name}")
   else:
       print("\nNo VM data was collected.")
if __name__ == "__main__":
   get_azure_update_manager_data(config_file="conf.json")