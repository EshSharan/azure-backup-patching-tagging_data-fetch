import json
import logging
import os
import sys
from azure.identity import ClientSecretCredential
from azure.mgmt.resource import ResourceManagementClient
from azure.mgmt.subscription import SubscriptionClient
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
# ------------------------ Logging Setup ------------------------
log_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "log")
os.makedirs(log_dir, exist_ok=True)
logging.basicConfig(
   filename=os.path.join(log_dir, "tagging.log"),
   level=logging.INFO,
   format='%(asctime)s - %(levelname)s - %(message)s',
   encoding='utf-8'
)
# ------------------------ Main Function ------------------------
def run_tagging(config_file="conf.json"):
   try:
       with open(config_file) as f:
           config = json.load(f)
       credential = ClientSecretCredential(
           tenant_id=config['tenantId'],
           client_id=config['clientId'],
           client_secret=config['clientSecret']
       )
       logging.info("🔑 Authentication successful!")
       print("🔑 Authentication successful!")
   except Exception as e:
       logging.critical(f"❌ Authentication failed: {e}")
       print(f"❌ Authentication failed: {e}")
       sys.exit(1)
   subscription_client = SubscriptionClient(credential)
   tagging_data = []
   for subscription in subscription_client.subscriptions.list():
       print(f"🌐 Processing subscription: {subscription.display_name}...")
       resource_client = ResourceManagementClient(credential, subscription.subscription_id)
       for resource in resource_client.resources.list():
           tags = resource.tags or {}
           existing_tags = "; ".join([f"{k}: {v}" for k, v in tags.items()]) if tags else "-"
           tagging_data.append({
               "Name": resource.name,
               "Resource Type": resource.type,
               "Resource Group": resource.id.split('/')[4] if len(resource.id.split('/')) > 4 else "-",
               "Location": resource.location,
               "Subscription": subscription.display_name,
               "Existing Tags": existing_tags,
               "Environment": tags.get("Environment", "-"),
               "Application": tags.get("Application", "-"),
               "Owner": tags.get("Owner", "-"),
               "Owner Email": tags.get("Owner Email", "-"),
               "Comments": tags.get("Comments", "-")
           })
   if tagging_data:
       df = pd.DataFrame(tagging_data)
       # Reorder columns: Existing Tags before individual tag columns
       cols = ["Name", "Resource Type", "Resource Group", "Location", "Subscription",
               "Existing Tags", "Environment", "Application", "Owner", "Owner Email", "Comments"]
       df = df[cols]
       output_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Tagging-Sheet.xlsx")
       df.to_excel(output_file, index=False)
       # ---------------- Table Styling ----------------
       wb = load_workbook(output_file)
       ws = wb.active
       thin_border = Border(
           left=Side(style='thin'),
           right=Side(style='thin'),
           top=Side(style='thin'),
           bottom=Side(style='thin')
       )
       # Header styling
       header_fill = PatternFill(start_color="FF00FF", end_color="FF00FF", fill_type="solid")
       header_font = Font(color="FFFFFF", bold=True, name="Calibri", size=12)
       for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=ws.max_column):
           for cell in col:
               cell.fill = header_fill
               cell.font = header_font
               cell.border = thin_border
               cell.alignment = Alignment(horizontal='center', vertical='center')
       # Data row styling
       data_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
       data_font = Font(color="000000", name="Calibri", size=12)
       for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
           for cell in row:
               cell.fill = data_fill
               cell.font = data_font
               cell.border = thin_border
       wb.save(output_file)
       logging.info(f"✅ Tagging report exported to {output_file} 🎉")
       print(f"✅ Tagging report exported to {output_file} 🎉")
   else:
       logging.info("⚠️ No resources found to export.")
       print("⚠️ No resources found to export.")
# ------------------------ Entry Point ------------------------
if __name__ == "__main__":
   print("✨ Running Tagging Script...")
   logging.info("✨ Running Tagging Script...")
   run_tagging()
   print("🌟 Tagging Script finished.")
   logging.info("🌟 Tagging Script finished.")